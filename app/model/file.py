from typing import Any
import openpyxl
import struct
import numpy as np
import gzip
import chardet as cd
import pandas as pd
import json as js
import pickle
import os


class Datas(object):
    '''
    Класс модели для работы с файлами.
    '''

    def __init__(self) -> None:
        pass

    @classmethod
    def load_txt(cls, filepath: str) -> pd.DataFrame:
        return pd.read_csv(
            filepath,
            sep=cls._get_sep(filepath),
            encoding=cls._get_enc(filepath),
            skiprows=[1]
        )

    @staticmethod
    def _get_sep(filepath: str) -> str:
        return '\t' if '.txt' in filepath else ','

    @staticmethod
    def _get_enc(filepath: str) -> str | None:
        with open(filepath, 'br') as input:
            return cd.detect(input.read(1000))['encoding']

    @staticmethod
    def write_xlsx(data: pd.DataFrame, plane_koef: dict, filepath: str) -> None:

        def find_bookmarks(search_value: str, sheet) -> tuple | None:
            search_value = '{' + search_value + '}'
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value == search_value:
                        return cell.row, cell.column
            return

        wb = openpyxl.load_workbook('templates/xls_template.xlsx')
        ws = wb.active

        if ws is None:
            raise ValueError('Неверный шаблон xlsx')

        for name in data.columns:
            address = find_bookmarks(name, ws)
            if not address:
                continue
            row, column = address
            for i, value in enumerate(data[name]):
                ws.cell(column=column + i, row=row, value=value) # type: ignore

        values = {}

        values['plane'] = plane_koef['name']
        values['pnk_kurs'] = plane_koef['values']['kurs_DISS_grad']
        values['pnk_kren'] = plane_koef['values']['kren_DISS_grad']
        values['pnk_tang'] = plane_koef['values']['tang_DISS_grad']

        values['mean_US'] = data.US.mean()
        values['mean_Wp'] = data.Wp.mean()
        values['average_US'] = np.average(data.US, weights=data.counts)
        values['average_Wp'] = np.average(data.Wp, weights=data.counts)
        values['std_US'] = data.US.std()
        values['std_Wp'] = data.Wp.std()
        values['skp_US'] = np.sqrt((data.US.sum() ** 2) / len(data.US))
        values['skp_Wp'] = np.sqrt((data.Wp.sum() ** 2) / len(data.Wp))
        values['skp_US2'] = values['skp_US'] * 2
        values['skp_Wp2'] = values['skp_Wp'] * 2
        values['abs_US'] = values['average_US'] + 2 * values['skp_US']
        values['abs_Wp'] = values['average_Wp'] + 2 * values['skp_Wp']

        for name, value in values.items():
            address = find_bookmarks(name, ws)
            if not address:
                continue
            row, column = address
            value = value if isinstance(value, str) else round(value, 3)
            ws.cell(column=column, row=row, value=value) # type: ignore

        wb.save(filepath)
        wb.close()

    @staticmethod
    def write_gzip(data: dict, filepath: str) -> None:
        with gzip.open(filepath, 'wb', compresslevel=1) as f:
            pickle.dump(data, f)

    @staticmethod
    def write_csv(data: pd.DataFrame, filepath: str) -> None:
        data.to_csv(filepath, index=False, sep=',')

    @staticmethod
    def load_gzip(filepath: str) -> Any:
        with gzip.open(filepath, 'rb') as f:
            return pickle.load(f)

    @staticmethod
    def load_csv(filepath: str) -> pd.DataFrame:
        return pd.read_csv(filepath)

    @staticmethod
    def load_python_script(filepath: str) -> str:
        with open(filepath, 'r', encoding='utf8') as file:
            return file.read()

    @staticmethod
    def load_json(filepath: str) -> Any:
        with open(filepath, 'r', encoding='utf8') as file:
            return js.load(file)

    @staticmethod
    def save_python(filepath: str, data) -> None:
        with open(filepath, 'w', encoding='utf8') as file:
            file.write(data)

    @staticmethod
    def save_json(filepath: str, data) -> None:
        with open(filepath, 'w', encoding='utf-8') as file:
            js.dump(data, file, indent=4, ensure_ascii=False)

    @staticmethod
    def get_list_json_in_folder(dirpath: str) -> list:
        '''Метод получения списка всех json файлов в папке'''
        result = []
        for item in os.scandir(dirpath):
            if item.is_file() and '.json' in item.name:
                result.append(item.path)
        return result

    @classmethod
    def get_jsons_data(cls, list_json: list) -> list:
        '''Метод получения данных все json файлов'''
        return [cls.load_json(filepath) for filepath in list_json]

    @staticmethod
    def get_mask_shift_from_field(size: str) -> tuple:
        # получаем маску и смещение для побитового сравнения и получения данных
        start, stop = [int(i) for i in size.split(':')]
        number_mask = int('1' * (stop - start + 1), 2)
        shift = start
        return number_mask, shift

    @staticmethod
    def get_unpacked_data_list(filepath: str) -> np.ndarray:
        with open(filepath, 'rb') as file:
            string = file.read()
            dtype = {
                'names': ['time', 'cs', 'null', 'values'], 'formats': ['>u4', 'u2', '4b', '>16h'], 
            }
            unpacked_data_list = np.frombuffer(string, dtype=dtype)
            unpacked_data_list = unpacked_data_list[['time', 'cs', 'values']]
        return unpacked_data_list

    @classmethod
    def unpack_element(
        cls, 
        byteswap: bool, 
        field_data: dict, 
        data_source: np.ndarray
        ) -> np.ndarray:
        position = field_data['position']
        koef = field_data['koef']
        size = field_data['size']
        type = field_data['type']
        mask, shift = cls.get_mask_shift_from_field(size)
        data_column = data_source[:, position]
        # if type == 'uint16' or type == 'pre':
        #     data_column = data_column.astype(np.uint16)

        if type == 'int16':
            data_column = data_column.astype(np.int16)

        if not byteswap:
            data_column = data_column.byteswap()
        if mask != 65535 or shift != 0:
            masked_data = np.bitwise_and(data_column, (mask << shift))
            masked_data = np.right_shift(masked_data, shift) 
        else:
            masked_data = data_column
        if type == 'int8':
            masked_data = masked_data.astype(np.int8)
        if type == 'uint16':
            masked_data = masked_data.astype(np.uint16)
        if type == 'pre':
            masked_data = masked_data.astype(np.int8)
        if koef != 1:
            masked_data = masked_data.astype(np.float64) * koef

        return masked_data

    @classmethod
    def unpack_group(
        cls, 
        byteswap: bool, 
        group_info: dict, 
        data_source: np.ndarray, 
        result_data: dict) -> None:
        if type(group_info['fields']) == list:
            for field in group_info['fields']:
                column_name = field['name']
                column = cls.unpack_element(byteswap, field, data_source)
                result_data[column_name] = column
        else:
            mask_condition, shift_condition = cls.get_mask_shift_from_field(
                group_info['condition_bit'])
            condition_data = (
                data_source[:, group_info['condition_byte'] + 2]
                 & (mask_condition << shift_condition)
                 ) >> shift_condition
            for condition, fields in group_info['fields'].items():
                condition_mask = condition_data == int(condition)
                for field in fields:
                    column_name = field['name']
                    column = cls.unpack_element(byteswap, field, data_source)
                    column = np.where(condition_mask, column, 0)
                    result_data[column_name] = column

    @staticmethod
    def get_filtered_data_by_checksum(
        checksum: int, 
        source_data: np.ndarray
        ) -> np.ndarray:
        control_sum = source_data['cs']
        control_sum_mask = control_sum == checksum
        data_list = source_data[control_sum_mask]
        return data_list

    @staticmethod
    def get_time_list(koef: float, source: np.ndarray) -> np.ndarray:
        time_list = source['time']
        time_list = time_list.byteswap() * koef
        return time_list

    @classmethod
    def unpack_fields(cls, byteswap, fields, data_list, result_dict) -> dict:
        for field in fields:
            if 'group' in field.keys() and field['group'] == True:
                cls.unpack_group(byteswap, field, data_list, result_dict)
            else:
                column_name = field['name']
                column = cls.unpack_element(byteswap, field, data_list)
                result_dict[column_name] = column

    @classmethod
    def unpack_adr(cls, adr, unpacked_data_list) -> dict:
        checksum = int(adr['checksum'], base=16)
        data_list = cls.get_filtered_data_by_checksum(
            checksum, unpacked_data_list
        )
        df_dict = {}
        df_dict['time'] = cls.get_time_list(adr['time_koef'], data_list)
        cls.unpack_fields(
            adr['bytes_swap'], adr['fields'], data_list['values'], df_dict)
        return df_dict

    @classmethod
    def load_pdd(cls, filepath_pdd: str, json_data: list) -> dict:
        unpacked_data_list = cls.get_unpacked_data_list(filepath_pdd)
        result_dict = {}

        for adr in json_data:
            adr_data = cls.unpack_adr(adr, unpacked_data_list)
            result_dict[adr['adr_name']] = pd.DataFrame(adr_data)

        return result_dict
